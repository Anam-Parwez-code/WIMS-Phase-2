"""
Utility functions for report generation
"""
from datetime import datetime, timedelta
from decimal import Decimal
from django.db.models import Sum, Count, Avg, Q


def format_currency(amount):
    """Format amount as Indian currency"""
    if amount is None:
        return "₹0.00"
    return f"₹{amount:,.2f}"


def format_date(date_obj, format='%d/%m/%Y'):
    """Format date object to string"""
    if not date_obj:
        return "-"
    return date_obj.strftime(format)


def calculate_date_range(period_type='month'):
    """
    Calculate date range based on period type
    
    Args:
        period_type: 'today', 'week', 'month', 'quarter', 'year'
    
    Returns:
        tuple: (from_date, to_date)
    """
    today = datetime.now().date()
    
    if period_type == 'today':
        return today, today
    
    elif period_type == 'week':
        # Last 7 days
        from_date = today - timedelta(days=7)
        return from_date, today
    
    elif period_type == 'month':
        # Current month
        from_date = today.replace(day=1)
        return from_date, today
    
    elif period_type == 'quarter':
        # Current quarter
        quarter_month = ((today.month - 1) // 3) * 3 + 1
        from_date = today.replace(month=quarter_month, day=1)
        return from_date, today
    
    elif period_type == 'year':
        # Current financial year (April to March)
        if today.month >= 4:
            from_date = today.replace(month=4, day=1)
        else:
            from_date = today.replace(year=today.year-1, month=4, day=1)
        return from_date, today
    
    else:
        # Default to current month
        from_date = today.replace(day=1)
        return from_date, today


def get_fee_statistics(queryset):
    """
    Calculate comprehensive fee statistics from FeeGeneration queryset
    
    Args:
        queryset: FeeGeneration queryset
    
    Returns:
        dict: Statistics including total fee, collected, outstanding
    """
    from fee_details.models import FeeDeposit
    
    total_fee = queryset.aggregate(total=Sum('total_fee'))['total'] or Decimal('0.00')
    total_advance = queryset.aggregate(total=Sum('advance_amount'))['total'] or Decimal('0.00')
    
    # Calculate total collected through installments
    fee_gen_ids = queryset.values_list('id', flat=True)
    total_installments = FeeDeposit.objects.filter(
        installment__fee_generation_id__in=fee_gen_ids,
        is_active=True
    ).aggregate(total=Sum('paid_amount'))['total'] or Decimal('0.00')
    
    total_collected = total_advance + total_installments
    total_outstanding = total_fee - total_collected
    
    collection_percentage = (total_collected / total_fee * 100) if total_fee > 0 else 0
    
    return {
        'total_fee': total_fee,
        'total_advance': total_advance,
        'total_installments': total_installments,
        'total_collected': total_collected,
        'total_outstanding': total_outstanding,
        'collection_percentage': collection_percentage,
        'student_count': queryset.count(),
    }


def get_admission_statistics(queryset):
    """
    Calculate admission statistics
    
    Args:
        queryset: Admission queryset
    
    Returns:
        dict: Statistics
    """
    total = queryset.count()
    
    stats = {
        'total': total,
        'male': queryset.filter(gender='Male').count(),
        'female': queryset.filter(gender='Female').count(),
        'other': queryset.filter(gender='Other').count(),
        'admitted': queryset.filter(status='Admitted').count(),
        'cancelled': queryset.filter(status='Cancelled').count(),
    }
    
    # Calculate percentages
    if total > 0:
        stats['male_percentage'] = (stats['male'] / total * 100)
        stats['female_percentage'] = (stats['female'] / total * 100)
        stats['admitted_percentage'] = (stats['admitted'] / total * 100)
    else:
        stats['male_percentage'] = 0
        stats['female_percentage'] = 0
        stats['admitted_percentage'] = 0
    
    return stats


def get_course_wise_stats(admissions_queryset):
    """
    Get course-wise admission statistics
    
    Args:
        admissions_queryset: Admission queryset
    
    Returns:
        list: List of dicts with course stats
    """
    from course.models import Course
    
    course_stats = []
    
    # Get all courses
    courses = Course.objects.filter(is_active=True)
    
    for course in courses:
        count = admissions_queryset.filter(courses=course).count()
        
        if count > 0:
            male = admissions_queryset.filter(courses=course, gender='Male').count()
            female = admissions_queryset.filter(courses=course, gender='Female').count()
            
            course_stats.append({
                'course_code': course.course_code,
                'course_name': course.course_name,
                'total': count,
                'male': male,
                'female': female,
            })
    
    # Sort by total (descending)
    course_stats.sort(key=lambda x: x['total'], reverse=True)
    
    return course_stats


def validate_date_range(from_date, to_date):
    """
    Validate and parse date range
    
    Args:
        from_date: String date in format YYYY-MM-DD
        to_date: String date in format YYYY-MM-DD
    
    Returns:
        tuple: (from_date_obj, to_date_obj) or raises ValueError
    """
    try:
        from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
        
        if from_date_obj > to_date_obj:
            raise ValueError("From date cannot be greater than to date")
        
        # Check if date range is not too large (e.g., max 1 year)
        if (to_date_obj - from_date_obj).days > 365:
            raise ValueError("Date range cannot exceed 1 year")
        
        return from_date_obj, to_date_obj
    
    except ValueError as e:
        raise ValueError(f"Invalid date format or range: {str(e)}")


def get_payment_mode_breakdown(fee_deposits_queryset):
    """
    Get payment mode wise breakdown
    
    Args:
        fee_deposits_queryset: FeeDeposit queryset
    
    Returns:
        list: List of dicts with payment mode stats
    """
    breakdown = fee_deposits_queryset.values('payment_mode').annotate(
        count=Count('id'),
        total_amount=Sum('paid_amount')
    ).order_by('-total_amount')
    
    return list(breakdown)


def get_daily_collection_breakdown(fee_deposits_queryset):
    """
    Get daily collection breakdown
    
    Args:
        fee_deposits_queryset: FeeDeposit queryset
    
    Returns:
        list: List of dicts with daily stats
    """
    breakdown = fee_deposits_queryset.values('payment_date').annotate(
        count=Count('id'),
        total_amount=Sum('paid_amount')
    ).order_by('payment_date')
    
    return list(breakdown)


def export_to_excel(data, filename, sheet_name='Sheet1'):
    """
    Export data to Excel file
    
    Args:
        data: 2D list with headers in first row
        filename: Output filename
        sheet_name: Excel sheet name
    
    Returns:
        BytesIO object containing Excel file
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Add data
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # Style header row
                if row_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install it using: pip install openpyxl")


class ReportFilters:
    """
    Common filters for reports
    """
    
    @staticmethod
    def apply_organization_filter(queryset, organization_id, org_field='organization'):
        """Apply organization filter to queryset"""
        if organization_id:
            filter_kwargs = {f'{org_field}_id': organization_id}
            return queryset.filter(**filter_kwargs)
        return queryset
    
    @staticmethod
    def apply_date_range_filter(queryset, from_date, to_date, date_field='created_at'):
        """Apply date range filter to queryset"""
        if from_date and to_date:
            filter_kwargs = {f'{date_field}__range': [from_date, to_date]}
            return queryset.filter(**filter_kwargs)
        return queryset
    
    @staticmethod
    def apply_status_filter(queryset, status, status_field='status'):
        """Apply status filter to queryset"""
        if status:
            filter_kwargs = {status_field: status}
            return queryset.filter(**filter_kwargs)
        return queryset
    
    @staticmethod
    def apply_active_filter(queryset, is_active=True):
        """Apply is_active filter to queryset"""
        return queryset.filter(is_active=is_active)