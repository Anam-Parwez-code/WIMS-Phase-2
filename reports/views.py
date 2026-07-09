from django.shortcuts import render
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from rest_framework import generics
from rest_framework import status
from django.db.models import Sum, Count, Q, F, Avg
from datetime import datetime, timedelta
from django.utils.dateparse import parse_date
from decimal import Decimal
from admission.models import Admission, CertificateApproval
from fee_details.models import FeeGeneration, FeeInstallment, FeeDeposit
from student_details.models import Registration
from course.models import Course
from master.models import Organization
from .pdf_base import BaseReportPDF
from datetime import datetime, timedelta
from decimal import Decimal
from reportlab.lib.units import inch 
from reportlab.platypus import Spacer
from admission.models import Admission, CertificateApproval
from rest_framework.permissions import IsAuthenticated
from master.models import Branch
from student_details.models import Enquiry, EnquiryFollowUp
from django.db.models import Exists, OuterRef
from django.db import models
from core.helper_function import get_branch_id
from staff.models import Attendance, Employee
from course.models import CourseTracker
from .serializers import AttendanceReportSerializer, CourseTrackerReportSerializer

from openpyxl import Workbook
from openpyxl.styles import Font
from django.http import HttpResponse

from django.http import HttpResponse

from reportlab.lib import colors

from reportlab.lib.styles import getSampleStyleSheet

from reportlab.platypus import Table

from reportlab.platypus import TableStyle

from reportlab.platypus import SimpleDocTemplate

from reportlab.platypus import Paragraph
from reportlab.lib.pagesizes import letter, landscape


def validate_course_wise_admission_filters(data):

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")
    course_id = data.get("courseId")

    if organization_id:

        if not Organization.objects.filter(
            id=organization_id,
            is_active=True
        ).exists():

            raise ValidationError({
                "organizationId":
                    "Selected organization does not exist."
            })

    if branch_id:

        queryset = Branch.objects.filter(
            id=branch_id,
            is_active=True
        )

        if organization_id:

            queryset = queryset.filter(
                organization_id=organization_id
            )

        if not queryset.exists():

            raise ValidationError({
                "branch":
                    "Selected branch does not belong to the selected organization."
            })

    if course_id:

        if not Course.objects.filter(
            id=course_id,
            is_active=True
        ).exists():

            raise ValidationError({
                "courseId":
                    "Selected course does not exist."
            })

def get_course_wise_admission_filters(request):

    data = request.data

    from_date = data.get("fromDate")
    to_date = data.get("toDate")

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")
    course_id = data.get("courseId")

    if not from_date or not to_date:

        raise ValidationError({
            "error":
                "fromDate and toDate are required."
        })

    validate_course_wise_admission_filters(data)

    courses = Course.objects.filter(
        is_active=True
    )

    if course_id:

        courses = courses.filter(
            id=course_id
        )

    return (
        courses,
        from_date,
        to_date,
        organization_id,
        branch_id
    )

class CourseWiseAdmissionReportAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        (
            courses,
            from_date,
            to_date,
            organization_id,
            branch_id,
        ) = get_course_wise_admission_filters(request)

        course_data = []

        total_admissions = 0
        total_male = 0
        total_female = 0

        for course in courses:

            admissions = Admission.objects.filter(
                course_batches__course=course,
                admission_date__range=[from_date, to_date],
                is_active=True,
                status="Admitted"
            ).select_related(
                "organization",
                "branch"
            ).distinct()

            if organization_id:
                admissions = admissions.filter(
                    organization_id=organization_id
                )

            if branch_id:
                admissions = admissions.filter(
                    branch_id=branch_id
                )

            count = admissions.count()

            if count > 0:

                male_count = admissions.filter(
                    gender='Male'
                ).count()

                female_count = admissions.filter(
                    gender='Female'
                ).count()

                course_data.append({

                    'course_id': course.id,

                    'course_code': course.course_code,

                    'course_name': course.course_name,

                    'male': male_count,

                    'female': female_count,

                    'total': count,
                })

                total_admissions += count
                total_male += male_count
                total_female += female_count

        course_data.sort(
            key=lambda x: x['total'],
            reverse=True
        )

        return Response({

            "report_metadata": {

                "report_name": "Course Wise Admission Report",

                "from_date": from_date,

                "to_date": to_date,

                "organization_id": organization_id,

                "branch_id": branch_id,

                "generated_at": datetime.now().isoformat()
            },

            "summary": {

                "total_courses_with_data": len(course_data),

                "grand_total_admissions": total_admissions,

                "grand_total_male": total_male,

                "grand_total_female": total_female
            },

            "results": course_data

        }, status=status.HTTP_200_OK)

class ExportCourseWiseAdmissionExcelAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        (
            courses,
            from_date,
            to_date,
            organization_id,
            branch_id
        ) = get_course_wise_admission_filters(request)

        wb = Workbook()

        ws = wb.active

        ws.title = "Course Wise Admission Report"

        headers = [

            "Course Code",

            "Course Name",

            "Male",

            "Female",

            "Total"

        ]

        for col, header in enumerate(headers, 1):

            cell = ws.cell(row=1, column=col)

            cell.value = header

            cell.font = Font(bold=True)

        row = 2

        for course in courses:

            admissions = Admission.objects.filter(
                course_batches__course=course,
                admission_date__range=[from_date, to_date],
                status="Admitted",
                is_active=True
            ).distinct()

            if organization_id:

                admissions = admissions.filter(
                    organization_id=organization_id
                )

            if branch_id:

                admissions = admissions.filter(
                    branch_id=branch_id
                )

            count = admissions.count()

            if count == 0:
                continue

            ws.cell(row=row, column=1).value = course.course_code

            ws.cell(row=row, column=2).value = course.course_name

            ws.cell(row=row, column=3).value = admissions.filter(
                gender="Male"
            ).count()

            ws.cell(row=row, column=4).value = admissions.filter(
                gender="Female"
            ).count()

            ws.cell(row=row, column=5).value = count

            row += 1

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = (
            'attachment; filename="Course_Wise_Admission_Report.xlsx"'
        )

        wb.save(response)

        return response

class ExportCourseWiseAdmissionPDFAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        (
            courses,
            from_date,
            to_date,
            organization_id,
            branch_id
        ) = get_course_wise_admission_filters(request)

        response = HttpResponse(
            content_type="application/pdf"
        )

        response["Content-Disposition"] = (
            'attachment; filename="Course_Wise_Admission_Report.pdf"'
        )

        doc = SimpleDocTemplate(response)

        styles = getSampleStyleSheet()

        elements = []

        elements.append(
            Paragraph(
                "<b>Course Wise Admission Report</b>",
                styles["Title"]
            )
        )

        table_data = [[

            "Course",

            "Male",

            "Female",

            "Total"

        ]]

        for course in courses:

            admissions = Admission.objects.filter(
                course_batches__course=course,
                admission_date__range=[from_date, to_date],
                status="Admitted",
                is_active=True
            ).distinct()

            if organization_id:

                admissions = admissions.filter(
                    organization_id=organization_id
                )

            if branch_id:

                admissions = admissions.filter(
                    branch_id=branch_id
                )

            count = admissions.count()

            if count == 0:
                continue

            table_data.append([

                course.course_name,

                admissions.filter(
                    gender="Male"
                ).count(),

                admissions.filter(
                    gender="Female"
                ).count(),

                count

            ])

        table = Table(table_data)

        table.setStyle(TableStyle([

            ("GRID", (0, 0), (-1, -1), 1, colors.black),

            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),

            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),

            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),

        ]))

        elements.append(table)

        doc.build(elements)

        return response



def validate_registration_report_filters(data):

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")

    if organization_id:

        if not Organization.objects.filter(
            id=organization_id,
            is_active=True
        ).exists():

            raise ValidationError({
                "organizationId": "Selected organization does not exist."
            })

    if branch_id:

        queryset = Branch.objects.filter(
            id=branch_id,
            is_active=True
        )

        if organization_id:

            queryset = queryset.filter(
                organization_id=organization_id
            )

        if not queryset.exists():

            raise ValidationError({
                "branch": "Selected branch does not belong to the selected organization."
            })
        
from django.db.models import Exists, OuterRef

def get_registration_report_queryset(request):

    data = request.data

    from_date = data.get("fromDate")
    to_date = data.get("toDate")

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")

    if not from_date or not to_date:

        raise ValidationError({
            "error": "fromDate and toDate are required."
        })

    validate_registration_report_filters(data)

    admission_exists = Admission.objects.filter(
        registration=OuterRef("pk"),
        is_active=True
    )

    queryset = Registration.objects.filter(
        registration_date__range=[from_date, to_date],
        is_active=True
    ).annotate(
        is_admitted=Exists(admission_exists)
    ).select_related(
        "organization",
        "branch"
    ).prefetch_related(
        "courses"
    )

    if organization_id:

        queryset = queryset.filter(
            organization_id=organization_id
        )

    if branch_id:

        queryset = queryset.filter(
            branch_id=branch_id
        )

    return queryset.order_by("-registration_date")

class RegistrationReportAPIView(APIView):
    """
    Generate registration report data in JSON format
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):

        try:

            data = request.data

            from_date = data.get("fromDate")
            to_date = data.get("toDate")
            organization_id = data.get("organizationId")
            branch_id = data.get("branch")

            if not from_date or not to_date:
                return Response(
                    {
                        "error": "fromDate and toDate are required"
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            # ============================================
            # CHECK ADMISSION EXISTS
            # ============================================

            admission_exists = Admission.objects.filter(
                registration=OuterRef("pk"),
                is_active=True
            )

            # ============================================
            # BASE QUERYSET
            # ============================================

            queryset = get_registration_report_queryset(request)

            # ============================================
            # FILTERS
            # ============================================

            # if organization_id:
            #     queryset = queryset.filter(
            #         organization_id=organization_id
            #     )

            # if branch_id:
            #     queryset = queryset.filter(
            #         branch_id=branch_id
            #     )

            # ============================================
            # SUMMARY
            # ============================================

            total_registrations = queryset.count()

            converted_to_admission = queryset.filter(
                is_admitted=True
            ).count()

            conversion_rate = 0

            if total_registrations > 0:
                conversion_rate = (
                    converted_to_admission
                    / total_registrations
                ) * 100

            # ============================================
            # RESULTS
            # ============================================

            registrations_list = []

            for reg in queryset:

                course_names = list(
                    reg.courses.values_list(
                        "course_name",
                        flat=True
                    )
                )

                registrations_list.append({

                    "id": reg.id,

                    "registration_code":
                        reg.registration_code,

                    "candidate_name":
                        reg.candidate_name,

                    "mobile_no":
                        reg.mobile_no,

                    "email":
                        reg.email or "-",

                    "gender":
                        reg.gender,

                    "registration_date":
                        reg.registration_date.strftime("%Y-%m-%d"),

                    "organization":
                        reg.organization.name
                        if reg.organization else None,

                    "branch":
                        reg.branch.name
                        if reg.branch else None,

                    "courses":
                        course_names,

                    "is_admitted":
                        reg.is_admitted
                })

            # ============================================
            # FINAL RESPONSE
            # ============================================

            return Response({

                "report_metadata": {

                    "report_name": "Registration Report",

                    "from_date": from_date,

                    "to_date": to_date,

                    "organization_id": organization_id,

                    "branch_id": branch_id,

                    "generated_at": datetime.now().isoformat()
                },

                "summary": {

                    "total_registrations":
                        total_registrations,

                    "converted_to_admission":
                        converted_to_admission,

                    "conversion_rate":
                        round(conversion_rate, 2)
                },

                "results":
                    registrations_list

            }, status=status.HTTP_200_OK)

        except Exception as e:

            return Response(
                {
                    "error": str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ExportRegistrationReportExcelAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_registration_report_queryset(request)

        wb = Workbook()

        ws = wb.active

        ws.title = "Registration Report"

        headers = [

            "Registration Code",

            "Student",

            "Mobile",

            "Email",

            "Gender",

            "Registration Date",

            "Organization",

            "Branch",

            "Courses",

            "Admitted"
        ]

        for col, header in enumerate(headers, 1):

            cell = ws.cell(row=1, column=col)

            cell.value = header

            cell.font = Font(bold=True)

        row = 2

        for reg in queryset:

            courses = ", ".join(
                reg.courses.values_list(
                    "course_name",
                    flat=True
                )
            )

            ws.cell(row=row, column=1).value = reg.registration_code

            ws.cell(row=row, column=2).value = reg.candidate_name

            ws.cell(row=row, column=3).value = reg.mobile_no

            ws.cell(row=row, column=4).value = reg.email

            ws.cell(row=row, column=5).value = reg.gender

            ws.cell(row=row, column=6).value = reg.registration_date.strftime("%d-%m-%Y")

            ws.cell(row=row, column=7).value = (
                reg.organization.name
                if reg.organization else ""
            )

            ws.cell(row=row, column=8).value = (
                reg.branch.name
                if reg.branch else ""
            )

            ws.cell(row=row, column=9).value = courses

            ws.cell(row=row, column=10).value = (
                "Yes"
                if reg.is_admitted
                else "No"
            )

            row += 1

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = (
            'attachment; filename="Registration_Report.xlsx"'
        )

        wb.save(response)

        return response

class ExportRegistrationReportPDFAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_registration_report_queryset(request)

        response = HttpResponse(
            content_type="application/pdf"
        )

        response["Content-Disposition"] = (
            'attachment; filename="Registration_Report.pdf"'
        )

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(letter)
        )

        styles = getSampleStyleSheet()

        elements = []

        elements.append(
            Paragraph(
                "<b>Registration Report</b>",
                styles["Title"]
            )
        )

        table_data = [[

            "Code",

            "Student",

            "Mobile",

            "Date",

            "Organization",

            "Branch",

            "Admitted"
        ]]

        for reg in queryset:

            table_data.append([

                reg.registration_code,

                reg.candidate_name,

                reg.mobile_no,

                reg.registration_date.strftime("%d-%m-%Y"),

                reg.organization.name
                if reg.organization else "",

                reg.branch.name
                if reg.branch else "",

                "Yes"
                if reg.is_admitted else "No"

            ])

        table = Table(table_data)

        table.setStyle(TableStyle([

            ("GRID", (0, 0), (-1, -1), 1, colors.black),

            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),

            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),

            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),

            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),

        ]))

        elements.append(table)

        doc.build(elements)

        return response







def validate_certificate_approval_filters(data):

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")

    # --------------------------------------
    # Validate Organization
    # --------------------------------------

    if organization_id:

        if not Organization.objects.filter(
            id=organization_id,
            is_active=True
        ).exists():

            raise ValidationError({
                "organizationId":
                    "Selected organization does not exist."
            })

    # --------------------------------------
    # Validate Branch
    # --------------------------------------

    if branch_id:

        queryset = Branch.objects.filter(
            id=branch_id,
            is_active=True
        )

        if organization_id:

            queryset = queryset.filter(
                organization_id=organization_id
            )

        if not queryset.exists():

            raise ValidationError({
                "branch":
                    "Selected branch does not belong to the selected organization."
            })

def get_certificate_approval_queryset(request):

    data = request.data

    from_date = data.get("fromDate")
    to_date = data.get("toDate")

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")

    if not from_date or not to_date:

        raise ValidationError({
            "error":
                "fromDate and toDate are required."
        })

    validate_certificate_approval_filters(data)

    queryset = CertificateApproval.objects.filter(
        approval_date__range=[from_date, to_date],
        is_active=True
    ).select_related(
        "admission",
        "admission__organization",
        "admission__branch",
        "approved_by"
    ).order_by("-approval_date")

    if organization_id:

        queryset = queryset.filter(
            admission__organization_id=organization_id
        )

    if branch_id:

        queryset = queryset.filter(
            admission__branch_id=branch_id
        )

    return queryset

class CertificateApprovalReportAPIView(APIView):
    """
    Generate report of certificate approvals in JSON format
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):

        try:

            data = request.data

            from_date = data.get("fromDate")
            to_date = data.get("toDate")
            # organization_id = data.get("organizationId")
            # branch_id = data.get("branchId")

            # ============================================
            # VALIDATION
            # ============================================

            if not from_date or not to_date:

                return Response(
                    {
                        "error":
                            "fromDate and toDate are required"
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            # ============================================
            # QUERYSET
            # ============================================

            # queryset = CertificateApproval.objects.filter(
            #     approval_date__range=[from_date, to_date],
            #     is_active=True
            # ).select_related(
            #     "admission",
            #     "admission__organization",
            #     "admission__branch",
            #     "approved_by"
            # ).order_by("-approval_date")

            queryset = get_certificate_approval_queryset(request)

            # ============================================
            # FILTERS
            # ============================================

            # if organization_id:

            #     queryset = queryset.filter(
            #         admission__organization_id=organization_id
            #     )

            # if branch_id:

            #     queryset = queryset.filter(
            #         admission__branch_id=branch_id
            #     )

            # ============================================
            # RESULTS
            # ============================================

            approvals_list = []

            for cert in queryset:

                approved_by_name = "N/A"

                if cert.approved_by:

                    approved_by_name = (
                        cert.approved_by.first_name
                        or cert.approved_by.username
                    )

                approvals_list.append({

                    "id":
                        cert.id,

                    "admission_no":
                        cert.admission.admission_code
                        if cert.admission else None,

                    "student_name":
                        cert.admission.candidate_name
                        if cert.admission else None,

                    "organization":
                        cert.admission.organization.name
                        if cert.admission and cert.admission.organization
                        else None,

                    "branch":
                        cert.admission.branch.name
                        if cert.admission and cert.admission.branch
                        else None,

                    "approval_date":
                        cert.approval_date.strftime("%Y-%m-%d")
                        if cert.approval_date else None,

                    "approved_by":
                        approved_by_name,

                    "remarks":
                        cert.remarks or ""
                })

            # ============================================
            # RESPONSE
            # ============================================

            return Response({

                "metadata": {

                    "report_name":
                        "Certificate Approval Report",

                    "from_date":
                        from_date,

                    "to_date":
                        to_date,

                    "generated_at":
                        datetime.now().isoformat()
                },

                "summary": {

                    "total_certificates_approved":
                        queryset.count()
                },

                "results":
                    approvals_list

            }, status=status.HTTP_200_OK)

        except Exception as e:

            return Response(
                {
                    "error": str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ExportCertificateApprovalExcelAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_certificate_approval_queryset(request)

        wb = Workbook()

        ws = wb.active

        ws.title = "Certificate Approval Report"

        headers = [

            "Approval Date",

            "Admission No",

            "Student Name",

            "Organization",

            "Branch",

            "Approved By",

            "Remarks"

        ]

        for col, header in enumerate(headers, 1):

            cell = ws.cell(row=1, column=col)

            cell.value = header

            cell.font = Font(bold=True)

        row = 2

        for cert in queryset:

            approved_by = ""

            if cert.approved_by:

                approved_by = (
                    cert.approved_by.first_name
                    or cert.approved_by.username
                )

            ws.cell(row=row, column=1).value = cert.approval_date.strftime("%d-%m-%Y")

            ws.cell(row=row, column=2).value = (
                cert.admission.admission_code
                if cert.admission else ""
            )

            ws.cell(row=row, column=3).value = (
                cert.admission.candidate_name
                if cert.admission else ""
            )

            ws.cell(row=row, column=4).value = (
                cert.admission.organization.name
                if cert.admission and cert.admission.organization
                else ""
            )

            ws.cell(row=row, column=5).value = (
                cert.admission.branch.name
                if cert.admission and cert.admission.branch
                else ""
            )

            ws.cell(row=row, column=6).value = approved_by

            ws.cell(row=row, column=7).value = cert.remarks or ""

            row += 1

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = (
            'attachment; filename="Certificate_Approval_Report.xlsx"'
        )

        wb.save(response)

        return response
    
class ExportCertificateApprovalPDFAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_certificate_approval_queryset(request)

        response = HttpResponse(
            content_type="application/pdf"
        )

        response["Content-Disposition"] = (
            'attachment; filename="Certificate_Approval_Report.pdf"'
        )

        doc = SimpleDocTemplate(response)

        styles = getSampleStyleSheet()

        elements = []

        elements.append(
            Paragraph(
                "<b>Certificate Approval Report</b>",
                styles["Title"]
            )
        )

        table_data = [[

            "Date",

            "Admission",

            "Student",

            "Organization",

            "Branch",

            "Approved By"

        ]]

        for cert in queryset:

            approved_by = ""

            if cert.approved_by:

                approved_by = (
                    cert.approved_by.first_name
                    or cert.approved_by.username
                )

            table_data.append([

                cert.approval_date.strftime("%d-%m-%Y"),

                cert.admission.admission_code
                if cert.admission else "",

                cert.admission.candidate_name
                if cert.admission else "",

                cert.admission.organization.name
                if cert.admission and cert.admission.organization
                else "",

                cert.admission.branch.name
                if cert.admission and cert.admission.branch
                else "",

                approved_by

            ])

        table = Table(table_data)

        table.setStyle(TableStyle([

            ("GRID", (0, 0), (-1, -1), 1, colors.black),

            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),

            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),

            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),

        ]))

        elements.append(table)

        doc.build(elements)

        return response




from datetime import datetime

from django.db import models
from django.db.models import Count, Q
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from admission.models import Admission
from course.models import Course

def get_filtered_queryset(request):

    data = request.data

    from_date = data.get("fromDate")
    to_date = data.get("toDate")

    admission_no = data.get("admissionNo")
    organization_id = data.get("organizationId")
    status_filter = data.get("status")
    gender_filter = data.get("gender")

    branch_id = data.get("branch")

    # branch_id = get_branch_id(request)

    # if not branch_id:
    #     branch_id = data.get("branch")

    # ============================================
    # VALIDATE ORGANIZATION
    # ============================================

    if organization_id:

        if not Organization.objects.filter(
            id=organization_id,
            is_active=True
        ).exists():

            raise ValidationError({
                "organizationId": "Selected organization does not exist."
            })

    # ============================================
    # VALIDATE BRANCH
    # ============================================

    if branch_id:

        branch = Branch.objects.filter(
            id=branch_id,
            is_active=True
        ).first()

        if not branch:

            raise ValidationError({
                "branch": "Selected branch does not exist."
            })

        if organization_id and branch.organization_id != int(organization_id):

            raise ValidationError({
                "branch": "Selected branch does not belong to the selected organization."
            })


    queryset = Admission.objects.filter(
        admission_date__range=[from_date, to_date],
        is_active=True
    )

    if admission_no:
        queryset = queryset.filter(
            admission_code__icontains=admission_no
        )

    if organization_id:
        queryset = queryset.filter(
            organization_id=organization_id
        )

    if branch_id:
        queryset = queryset.filter(
            branch_id=branch_id
        )

    if status_filter:
        queryset = queryset.filter(
            status=status_filter
        )

    if gender_filter:
        queryset = queryset.filter(
            gender=gender_filter
        )

    return queryset.select_related(
        "organization",
        "branch"
    ).prefetch_related(
        "course_batches__course",
        "course_batches__batch"
    ).distinct()

class GenerateAdmissionReportAPIView(APIView):
    """
    Generate detailed admission report statistics and data
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):

        try:

            data = request.data

            from_date = data.get("fromDate")
            to_date = data.get("toDate")

            admission_no = data.get("admissionNo")
            organization_id = data.get("organizationId")

            status_filter = data.get("status")
            gender_filter = data.get("gender")

            branch_id = data.get("branch")

            # branch_id = get_branch_id(request)

            # if not branch_id:
            #     branch_id = data.get("branch")

            # ============================================
            # VALIDATE ORGANIZATION
            # ============================================

            if organization_id:

                if not Organization.objects.filter(
                    id=organization_id,
                    is_active=True
                ).exists():

                    return Response(
                        {
                            "error": "Selected organization does not exist."
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # ============================================
            # VALIDATE BRANCH
            # ============================================

            if branch_id:

                branch = Branch.objects.filter(
                    id=branch_id,
                    is_active=True
                ).first()

                if not branch:

                    return Response(
                        {
                            "error": "Selected branch does not exist."
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )

                if organization_id and branch.organization_id != int(organization_id):

                    return Response(
                        {
                            "error": "Selected branch does not belong to the selected organization."
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # ============================================
            # VALIDATION
            # ============================================

            if not from_date or not to_date:

                return Response(
                    {
                        "error":
                            "fromDate and toDate are required"
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            # ============================================
            # BASE QUERYSET
            # ============================================

            queryset = Admission.objects.filter(
                admission_date__range=[
                    from_date,
                    to_date
                ],
                is_active=True
            )

            # ============================================
            # FILTERS
            # ============================================

            if admission_no:

                queryset = queryset.filter(
                    admission_code__icontains=admission_no
                )

            if organization_id:

                queryset = queryset.filter(
                    organization_id=organization_id
                )

            if branch_id:

                queryset = queryset.filter(
                    branch_id=branch_id
                )

            if status_filter:

                queryset = queryset.filter(
                    status=status_filter
                )

            if gender_filter:

                queryset = queryset.filter(
                    gender=gender_filter
                )

            # ============================================
            # OPTIMIZATION
            # ============================================

            queryset = queryset.select_related(
                "organization",
                "branch"
            ).prefetch_related(
                "course_batches__course",
                "course_batches__batch"
            ).distinct()

            # ============================================
            # SUMMARY
            # ============================================

            total_admissions = queryset.count()

            stats = queryset.aggregate(

                male=Count(
                    "id",
                    filter=Q(gender="Male")
                ),

                female=Count(
                    "id",
                    filter=Q(gender="Female")
                ),

                admitted=Count(
                    "id",
                    filter=Q(status="Admitted")
                ),

                cancelled=Count(
                    "id",
                    filter=Q(status="Cancelled")
                )
            )

            # ============================================
            # COURSE WISE BREAKDOWN
            # ============================================

            course_breakdown = (

                queryset.values(
                    "course_batches__course__id",
                    "course_batches__course__course_name",
                    "course_batches__course__course_code"
                )

                .annotate(
                    student_count=Count(
                        "id",
                        distinct=True
                    )
                )

                .exclude(
                    course_batches__course__id__isnull=True
                )

                .order_by("-student_count")
            )

            # ============================================
            # STUDENT LIST
            # ============================================

            admissions_list = []

            for adm in queryset:

                # ========================================
                # COURSE + BATCH GROUPING
                # ========================================

                course_map = {}

                for cb in adm.course_batches.all():

                    course_id = cb.course.id

                    if course_id not in course_map:

                        course_map[course_id] = {

                            "course_id": cb.course.id,

                            "course_name": cb.course.course_name,

                            "course_code": cb.course.course_code,

                            "batches": []
                        }

                    course_map[course_id]["batches"].append({

                        "batch_id": cb.batch.id,

                        "batch_name": cb.batch.batch_name,

                        "batch_code": cb.batch.batch_code
                    })

                admissions_list.append({

                    "id": adm.id,

                    "admission_code": adm.admission_code,

                    "student_name": adm.candidate_name,

                    "mobile_no": adm.mobile_no,

                    "alternate_mobile_no": adm.alternate_mobile_no,

                    "email": adm.email or "-",

                    "gender": adm.gender,

                    "admission_date": adm.admission_date.strftime(
                        "%Y-%m-%d"
                    ),

                    "father_name": adm.father_name,

                    "mother_name": adm.mother_name,

                    "qualification": adm.qualification,

                    "aadhaar_no": adm.aadhaar_no,

                    "address": adm.address,

                    "status": adm.status,

                    # ====================================
                    # ORGANIZATION
                    # ====================================

                    "organization": {

                        "id": (
                            adm.organization.id
                            if adm.organization else None
                        ),

                        "name": (
                            adm.organization.name
                            if adm.organization else None
                        )
                    },

                    # ====================================
                    # BRANCH
                    # ====================================

                    "branch": {

                        "id": (
                            adm.branch.id
                            if adm.branch else None
                        ),

                        "name": (
                            adm.branch.name
                            if adm.branch else None
                        )
                    },

                    # ====================================
                    # COURSE DETAILS
                    # ====================================

                    "courses": list(course_map.values())
                })

            # ============================================
            # RESPONSE
            # ============================================

            return Response({

                "metadata": {

                    "report_name": "Admission Report",

                    "from_date": from_date,

                    "to_date": to_date,

                    "generated_at": datetime.now().isoformat(),

                    "branch_id": branch_id,

                    "organization_id": organization_id
                },

                "summary": {

                    "total_admissions": total_admissions,

                    "gender_split": {

                        "male": stats["male"],

                        "female": stats["female"]
                    },

                    "status_split": {

                        "admitted": stats["admitted"],

                        "cancelled": stats["cancelled"]
                    },

                    "course_wise_breakdown": [

                        {
                            "course_id":
                                item[
                                    "course_batches__course__id"
                                ],

                            "course_name":
                                item[
                                    "course_batches__course__course_name"
                                ],

                            "course_code":
                                item[
                                    "course_batches__course__course_code"
                                ],

                            "student_count":
                                item["student_count"]
                        }

                        for item in course_breakdown
                    ]
                },

                "results": admissions_list

            }, status=status.HTTP_200_OK)

        except Exception as e:

            return Response(
                {
                    "error": str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ExportAdmissionReportExcelAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_filtered_queryset(request)

        wb = Workbook()

        ws = wb.active

        ws.title = "Admission Report"

        headers = [

            "Admission No",
            "Student Name",
            "Gender",
            "Mobile",
            "Organization",
            "Branch",
            "Admission Date",
            "Status",
            "Course",
            "Batch"
        ]

        for col, value in enumerate(headers, 1):

            cell = ws.cell(row=1, column=col)

            cell.value = value

            cell.font = Font(bold=True)

        row = 2

        for admission in queryset:

            courses = []

            batches = []

            for cb in admission.course_batches.all():

                courses.append(cb.course.course_name)

                batches.append(cb.batch.batch_name)

            ws.cell(row=row, column=1).value = admission.admission_code

            ws.cell(row=row, column=2).value = admission.candidate_name

            ws.cell(row=row, column=3).value = admission.gender

            ws.cell(row=row, column=4).value = admission.mobile_no

            ws.cell(row=row, column=5).value = (
                admission.organization.name
                if admission.organization else ""
            )

            ws.cell(row=row, column=6).value = (
                admission.branch.name
                if admission.branch else ""
            )

            ws.cell(row=row, column=7).value = admission.admission_date.strftime("%d-%m-%Y")

            ws.cell(row=row, column=8).value = admission.status

            ws.cell(row=row, column=9).value = ", ".join(courses)

            ws.cell(row=row, column=10).value = ", ".join(batches)

            row += 1

        response = HttpResponse(

            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        )

        response["Content-Disposition"] = 'attachment; filename="Admission_Report.xlsx"'

        wb.save(response)

        return response

class ExportAdmissionReportPDFAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_filtered_queryset(request)

        response = HttpResponse(content_type="application/pdf")

        response["Content-Disposition"] = 'attachment; filename="Admission_Report.pdf"'

        doc = SimpleDocTemplate(response)

        styles = getSampleStyleSheet()

        elements = []

        elements.append(
            Paragraph(
                "<b>Admission Report</b>",
                styles["Title"]
            )
        )

        data = [[

            "Admission No",

            "Student",

            "Gender",

            "Organization",

            "Branch",

            "Status"
        ]]

        for admission in queryset:

            data.append([

                admission.admission_code,

                admission.candidate_name,

                admission.gender,

                admission.organization.name if admission.organization else "",

                admission.branch.name if admission.branch else "",

                admission.status
            ])

        table = Table(data)

        table.setStyle(TableStyle([

            ("BACKGROUND", (0,0), (-1,0), colors.grey),

            ("TEXTCOLOR",(0,0),(-1,0),colors.whitesmoke),

            ("GRID",(0,0),(-1,-1),1,colors.black),

            ("BACKGROUND",(0,1),(-1,-1),colors.beige),

            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

            ("BOTTOMPADDING",(0,0),(-1,0),10)

        ]))

        elements.append(table)

        doc.build(elements)

        return response



def validate_fee_report_filters(data):

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")

    # ----------------------------------
    # Validate Organization
    # ----------------------------------

    if organization_id:

        if not Organization.objects.filter(
            id=organization_id,
            is_active=True
        ).exists():

            raise ValidationError({
                "organizationId":
                    "Selected organization does not exist."
            })

    # ----------------------------------
    # Validate Branch
    # ----------------------------------

    if branch_id:

        branch_queryset = Branch.objects.filter(
            id=branch_id,
            is_active=True
        )

        if organization_id:

            branch_queryset = branch_queryset.filter(
                organization_id=organization_id
            )

        if not branch_queryset.exists():

            raise ValidationError({
                "branch":
                    "Invalid branch or the selected branch does not belong to the selected organization."
            })

def get_fee_report_queryset(request):

    data = request.data

    from_date = data.get("fromDate")
    to_date = data.get("toDate")

    admission_no = data.get("admissionNo")
    payment_mode_id = data.get("paymentMode")

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")

    validate_fee_report_filters(data)

    queryset = FeeDeposit.objects.filter(
        payment_date__range=[from_date, to_date],
        is_active=True
    ).select_related(
        "payment_mode",
        "installment__fee_generation__admission"
    ).order_by("payment_date")

    if admission_no:

        queryset = queryset.filter(
            installment__fee_generation__admission__admission_code=admission_no
        )

    if payment_mode_id:

        queryset = queryset.filter(
            payment_mode_id=payment_mode_id
        )

    if organization_id:

        queryset = queryset.filter(
            installment__fee_generation__admission__organization_id=organization_id
        )

    if branch_id:

        queryset = queryset.filter(
            installment__fee_generation__admission__branch_id=branch_id
        )

    return queryset

class FeeDepositReportAPIView(APIView):
    """
    Generate detailed fee deposit/payment report in JSON format
    """
    
    def post(self, request):
        try:
            data = request.data
            from_date = data.get('fromDate')
            to_date = data.get('toDate')
            adm_no = data.get('admissionNo')
            payment_mode_id = data.get('paymentMode')  # Now expecting ID
            organization_id = data.get('organizationId')

            # 1. Validation
            if not from_date or not to_date:
                return Response(
                    {"error": "fromDate and toDate are required"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 2. Build Queryset
            # Using select_related to navigate the deep relationship: Deposit -> Installment -> FeeGen -> Admission
            queryset = get_fee_report_queryset(request)


            # 3. Aggregated Summary Statistics
            summary_stats = queryset.aggregate(
                total_count=Count('id'),
                total_collected=Sum('paid_amount')
            )
            
            # Payment mode breakdown (Now grouping by the related PaymentMethod name)
            mode_breakdown = queryset.values('payment_mode__name').annotate(
                count=Count('id'),
                total=Sum('paid_amount')
            ).order_by('-total')

            # 4. Format Detailed Results
            results = []
            for dep in queryset:
                # Safe navigation for related objects
                inst = dep.installment
                fee_gen = inst.fee_generation if inst else None
                adm = fee_gen.admission if fee_gen else None
                
                results.append({
                    "id": dep.id,
                    "payment_date": dep.payment_date.strftime('%Y-%m-%d'),
                    "admission_no": adm.admission_code if adm else "N/A",
                    "student_name": adm.candidate_name if adm else "N/A",
                    "installment_no": inst.installment_no if inst else "N/A",
                    "payment_mode": dep.payment_mode.name if dep.payment_mode else "N/A",
                    "paid_amount": float(dep.paid_amount),
                    "reference_no": dep.reference_no or "-",
                    "bank_name": dep.bank_name or "-",
                    "organization": {
                        "id": adm.organization.id if adm and adm.organization else None,
                        "name": adm.organization.name if adm and adm.organization else None
                    },

                    "branch": {
                        "id": adm.branch.id if adm and adm.branch else None,
                        "name": adm.branch.name if adm and adm.branch else None
                    }
                })

            # 5. Return JSON Response
            return Response({
                "metadata": {
                    "report_name": "Fee Collection Report",
                    "from_date": from_date,
                    "to_date": to_date,
                    "generated_at": datetime.now().isoformat()
                },
                "summary": {
                    "total_deposits": summary_stats['total_count'] or 0,
                    "total_amount_collected": float(summary_stats['total_collected'] or 0),
                    "mode_wise_breakdown": [
                        {
                            "mode": item['payment_mode__name'],
                            "count": item['count'],
                            "total_amount": float(item['total'])
                        } for item in mode_breakdown
                    ]
                },
                "results": results
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ExportFeeDepositReportExcelAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_fee_report_queryset(request)

        wb = Workbook()

        ws = wb.active

        ws.title = "Fee Collection Report"

        headers = [

            "Payment Date",

            "Admission No",

            "Student Name",

            "Installment",

            "Payment Mode",

            "Amount",

            "Reference No",

            "Bank",

            "Organization",

            "Branch"
        ]

        for col, header in enumerate(headers, 1):

            cell = ws.cell(row=1, column=col)

            cell.value = header

            cell.font = Font(bold=True)

        row = 2

        for dep in queryset:

            inst = dep.installment
            fee = inst.fee_generation if inst else None
            adm = fee.admission if fee else None

            ws.cell(row=row, column=1).value = dep.payment_date.strftime("%d-%m-%Y")

            ws.cell(row=row, column=2).value = (adm.admission_code if adm else "")

            ws.cell(row=row, column=3).value = (adm.candidate_name if adm else "")

            ws.cell(row=row, column=4).value = inst.installment_no

            ws.cell(row=row, column=5).value = dep.payment_mode.name if dep.payment_mode else ""

            ws.cell(row=row, column=6).value = float(dep.paid_amount)

            ws.cell(row=row, column=7).value = dep.reference_no

            ws.cell(row=row, column=8).value = dep.bank_name

            ws.cell(row=row, column=9).value = (
                adm.organization.name
                if adm and adm.organization else ""
            )

            ws.cell(row=row, column=10).value = (
                adm.branch.name
                if adm and adm.branch else ""
            )

            row += 1

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = 'attachment; filename="Fee_Collection_Report.xlsx"'

        wb.save(response)

        return response

class ExportFeeDepositReportPDFAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_fee_report_queryset(request)

        response = HttpResponse(content_type="application/pdf")

        response["Content-Disposition"] = 'attachment; filename="Fee_Collection_Report.pdf"'

        doc = SimpleDocTemplate(response)

        styles = getSampleStyleSheet()

        elements = []

        elements.append(
            Paragraph(
                "<b>Fee Collection Report</b>",
                styles["Title"]
            )
        )

        table_data = [[
            "Date",
            "Admission",
            "Student",
            "Organization",
            "Branch",
            "Mode",
            "Amount"
        ]]

        for dep in queryset:

            inst = dep.installment
            fee = inst.fee_generation if inst else None
            adm = fee.admission if fee else None

            table_data.append([

                dep.payment_date.strftime("%d-%m-%Y"),

                adm.admission_code if adm else "",

                adm.candidate_name if adm else "",

                adm.organization.name
                if adm and adm.organization else "",

                adm.branch.name
                if adm and adm.branch else "",

                dep.payment_mode.name
                if dep.payment_mode else "",

                str(dep.paid_amount)
            ])

        table = Table(table_data)

        table.setStyle(TableStyle([

            ("GRID", (0, 0), (-1, -1), 1, colors.black),

            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),

            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),

            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),

        ]))

        elements.append(table)

        doc.build(elements)

        return response





def validate_outstanding_fee_filters(data):

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")

    # ---------------------------------
    # Validate Organization
    # ---------------------------------

    if organization_id:

        if not Organization.objects.filter(
            id=organization_id,
            is_active=True
        ).exists():

            raise ValidationError({
                "organizationId":
                    "Selected organization does not exist."
            })

    # ---------------------------------
    # Validate Branch
    # ---------------------------------

    if branch_id:

        queryset = Branch.objects.filter(
            id=branch_id,
            is_active=True
        )

        if organization_id:

            queryset = queryset.filter(
                organization_id=organization_id
            )

        if not queryset.exists():

            raise ValidationError({
                "branch":
                    "Invalid branch or the selected branch does not belong to the selected organization."
            })

from decimal import Decimal
from django.utils.dateparse import parse_date

def get_outstanding_fee_queryset(request):

    data = request.data

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")

    min_outstanding = Decimal(
        str(data.get("minOutstanding", 0))
    )

    from_date = data.get("fromDate")
    to_date = data.get("toDate")

    validate_outstanding_fee_filters(data)

    queryset = FeeGeneration.objects.filter(
        is_active=True,
        admission__is_active=True,
        balance_amount__gte=min_outstanding
    ).select_related(
        "admission",
        "admission__organization",
        "admission__branch"
    ).order_by("-balance_amount")

    if from_date:

        queryset = queryset.filter(
            admission__admission_date__gte=parse_date(from_date)
        )

    if to_date:

        queryset = queryset.filter(
            admission__admission_date__lte=parse_date(to_date)
        )

    if organization_id:

        queryset = queryset.filter(
            admission__organization_id=organization_id
        )

    if branch_id:

        queryset = queryset.filter(
            admission__branch_id=branch_id
        )

    return queryset

class OutstandingFeeReportAPIView(APIView):
    """
    Generate JSON report of students with outstanding fees filtered by date range.
    """
    
    def post(self, request):
        try:
            data = request.data
            # organization_id = data.get('organizationId')
            # branch_id = get_branch_id(request)
            min_outstanding = Decimal(str(data.get('minOutstanding', 0)))
            
            # Extract Dates
            from_date = data.get('fromDate')
            to_date = data.get('toDate')

            # 1. Build Queryset
            # queryset = FeeGeneration.objects.filter(
            #     is_active=True,
            #     admission__is_active=True,
            #     balance_amount__gte=min_outstanding
            # ).select_related('admission', 'admission__organization').order_by('-balance_amount')
            
            queryset = get_outstanding_fee_queryset(request)

            # Apply Date Filters (Using admission_date as the primary timeline)
            # if from_date:
            #     queryset = queryset.filter(admission__admission_date__gte=parse_date(from_date))
            # if to_date:
            #     queryset = queryset.filter(admission__admission_date__lte=parse_date(to_date))
            
            # if organization_id:
            #     queryset = queryset.filter(admission__organization_id=organization_id)
            
            # if branch_id:
            #     queryset = queryset.filter(branch_id=branch_id)

            # 2. Aggregated Summary
            summary_stats = queryset.aggregate(
                grand_total_outstanding=Sum('balance_amount'),
                student_count=models.Count('id')
            )

            # 3. Format Results
            outstanding_list = []
            for fee_gen in queryset:
                adm = fee_gen.admission
                outstanding_list.append({

                    "id": fee_gen.id,

                    "admission_code": adm.admission_code,

                    "student_name": adm.candidate_name,

                    "mobile": adm.mobile_no,

                    "organization": {

                        "id": adm.organization.id if adm.organization else None,

                        "name": adm.organization.name if adm.organization else None
                    },

                    "branch": {

                        "id": adm.branch.id if adm.branch else None,

                        "name": adm.branch.name if adm.branch else None
                    },

                    "total_fee": float(fee_gen.total_fee),

                    "paid_amount": float(
                        fee_gen.total_fee - fee_gen.balance_amount
                    ),

                    "outstanding_amount": float(
                        fee_gen.balance_amount
                    ),

                    "installment_count": fee_gen.installment_count,

                    "last_updated": fee_gen.updated_at.strftime("%Y-%m-%d"),

                    "admission_date": adm.admission_date.strftime("%Y-%m-%d")
                })

            # 4. Final JSON Response
            return Response({
                "metadata": {
                    "report_name": "Outstanding Fee Report",
                    "from_date": from_date,
                    "to_date": to_date,
                    "min_outstanding_filter": float(min_outstanding),
                    "generated_at": datetime.now().isoformat()
                },
                "summary": {
                    "total_students_with_outstanding": summary_stats['student_count'] or 0,
                    "total_outstanding_amount": float(summary_stats['grand_total_outstanding'] or 0)
                },
                "results": outstanding_list
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ExportOutstandingFeeReportExcelAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_outstanding_fee_queryset(request)

        wb = Workbook()

        ws = wb.active

        ws.title = "Outstanding Fee Report"

        headers = [

            "Admission No",

            "Student Name",

            "Organization",

            "Branch",

            "Mobile",

            "Admission Date",

            "Total Fee",

            "Paid Amount",

            "Outstanding Amount",

            "Installments"

        ]

        for col, header in enumerate(headers, 1):

            cell = ws.cell(row=1, column=col)

            cell.value = header

            cell.font = Font(bold=True)

        row = 2

        for fee in queryset:

            adm = fee.admission

            ws.cell(row=row, column=1).value = adm.admission_code

            ws.cell(row=row, column=2).value = adm.candidate_name

            ws.cell(row=row, column=3).value = (
                adm.organization.name
                if adm.organization else ""
            )

            ws.cell(row=row, column=4).value = (
                adm.branch.name
                if adm.branch else ""
            )

            ws.cell(row=row, column=5).value = adm.mobile_no

            ws.cell(row=row, column=6).value = adm.admission_date.strftime("%d-%m-%Y")

            ws.cell(row=row, column=7).value = float(fee.total_fee)

            ws.cell(row=row, column=8).value = float(
                fee.total_fee - fee.balance_amount
            )

            ws.cell(row=row, column=9).value = float(
                fee.balance_amount
            )

            ws.cell(row=row, column=10).value = fee.installment_count

            row += 1

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = (
            'attachment; filename="Outstanding_Fee_Report.xlsx"'
        )

        wb.save(response)

        return response

class ExportOutstandingFeeReportPDFAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_outstanding_fee_queryset(request)

        response = HttpResponse(
            content_type="application/pdf"
        )

        response["Content-Disposition"] = (
            'attachment; filename="Outstanding_Fee_Report.pdf"'
        )

        doc = SimpleDocTemplate(response)

        styles = getSampleStyleSheet()

        elements = []

        elements.append(
            Paragraph(
                "<b>Outstanding Fee Report</b>",
                styles["Title"]
            )
        )

        table_data = [[

            "Admission",

            "Student",

            "Organization",

            "Branch",

            "Outstanding"

        ]]

        for fee in queryset:

            adm = fee.admission

            table_data.append([

                adm.admission_code,

                adm.candidate_name,

                adm.organization.name
                if adm.organization else "",

                adm.branch.name
                if adm.branch else "",

                str(fee.balance_amount)

            ])

        table = Table(table_data)

        table.setStyle(TableStyle([

            ("GRID", (0, 0), (-1, -1), 1, colors.black),

            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),

            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),

            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),

        ]))

        elements.append(table)

        doc.build(elements)

        return response







def validate_enquiry_report_filters(data):

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")

    if organization_id:

        if not Organization.objects.filter(
            id=organization_id,
            is_active=True
        ).exists():

            raise ValidationError({
                "organizationId":
                    "Selected organization does not exist."
            })

    if branch_id:

        queryset = Branch.objects.filter(
            id=branch_id,
            is_active=True
        )

        if organization_id:

            queryset = queryset.filter(
                organization_id=organization_id
            )

        if not queryset.exists():

            raise ValidationError({
                "branch":
                    "Selected branch does not belong to the selected organization."
            })

def get_enquiry_report_queryset(request):

    data = request.data

    from_date = data.get("fromDate")
    to_date = data.get("toDate")

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")

    if not from_date or not to_date:

        raise ValidationError({
            "error": "fromDate and toDate are required."
        })

    validate_enquiry_report_filters(data)

    queryset = Enquiry.objects.filter(
        enquiry_date__range=[from_date, to_date],
        is_active=True
    ).select_related(
        "assigned_to",
        "assigned_to__organization",
        "assigned_to__branch",
        "enquiry_source",
        "status",
        "followup_medium"
    ).prefetch_related(
        "courses"
    )

    if organization_id:

        queryset = queryset.filter(
            assigned_to__organization_id=organization_id
        )

    if branch_id:

        queryset = queryset.filter(
            assigned_to__branch_id=branch_id
        )

    return queryset

class EnquiryReportAPIView(APIView):
    """
    Generate JSON report for Enquiries stored in student_details
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            data = request.data
            from_date = data.get('fromDate')
            to_date = data.get('toDate')
            org_id = data.get('organizationId')

            if not from_date or not to_date:
                return Response({"error": "fromDate and toDate are required"}, status=400)

            # 1. Fetch Enquiries with related master data
            # queryset = Enquiry.objects.filter(
            #     enquiry_date__range=[from_date, to_date],
            #     is_active=True
            # ).select_related(
            #     'enquiry_source', 
            #     'status', 
            #     'assigned_to', 
            #     'followup_medium'
            # ).prefetch_related('courses')

            queryset = get_enquiry_report_queryset(request)

         

            # 2. Aggregated Statistics for Charts/Summary
            source_stats = queryset.values('enquiry_source__name').annotate(count=Count('id'))
            status_stats = queryset.values('status__name').annotate(count=Count('id'))

            # 3. Format result list
            results = []

            for enq in queryset:

                results.append({

                    "id": enq.id,

                    "enquiry_code": enq.enquiry_code,

                    "date": enq.enquiry_date.strftime("%Y-%m-%d"),

                    "candidate_name": enq.candidate_name,

                    "mobile": enq.mobile_no,

                    "organization": {
                        "id": (
                            enq.assigned_to.organization.id
                            if enq.assigned_to and enq.assigned_to.organization
                            else None
                        ),
                        "name": (
                            enq.assigned_to.organization.name
                            if enq.assigned_to and enq.assigned_to.organization
                            else None
                        )
                    },


                    "branch": {
                        "id": (
                            enq.assigned_to.branch.id
                            if enq.assigned_to and enq.assigned_to.branch
                            else None
                        ),
                        "name": (
                            enq.assigned_to.branch.name
                            if enq.assigned_to and enq.assigned_to.branch
                            else None
                        )
                    },

                    "source": (
                        enq.enquiry_source.name
                        if enq.enquiry_source else "Direct"
                    ),

                    "status": (
                        enq.status.name
                        if enq.status else "Open"
                    ),

                    "assigned_to": (
                        enq.assigned_to.name
                        if enq.assigned_to else "Unassigned"
                    ),

                    "courses": [
                        c.course_name
                        for c in enq.courses.all()
                    ],

                    "next_followup": (
                        enq.next_followup_date.strftime("%Y-%m-%d")
                        if enq.next_followup_date else None
                    )
                })

            return Response({
                "summary": {
                    "total_enquiries": queryset.count(),
                    "by_source": {item['enquiry_source__name']: item['count'] for item in source_stats if item['enquiry_source__name']},
                    "by_status": {item['status__name']: item['count'] for item in status_stats if item['status__name']}
                },
                "results": results
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({"error": str(e)}, status=500)

class ExportEnquiryReportExcelAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_enquiry_report_queryset(request)

        wb = Workbook()

        ws = wb.active

        ws.title = "Enquiry Report"

        headers = [

            "Date",

            "Enquiry No",

            "Candidate",

            "Mobile",

            "Organization",

            "Branch",

            "Source",

            "Status",

            "Assigned To",

            "Courses"

        ]

        for col, header in enumerate(headers, 1):

            cell = ws.cell(row=1, column=col)

            cell.value = header

            cell.font = Font(bold=True)

        row = 2

        for enq in queryset:

            ws.cell(row=row, column=1).value = enq.enquiry_date.strftime("%d-%m-%Y")

            ws.cell(row=row, column=2).value = enq.enquiry_code

            ws.cell(row=row, column=3).value = enq.candidate_name

            ws.cell(row=row, column=4).value = enq.mobile_no

            ws.cell(row=row, column=5).value = (
                enq.assigned_to.organization.name \
                if enq.assigned_to and enq.assigned_to.organization else ""
            )

            ws.cell(row=row, column=6).value = (
                enq.assigned_to.branch.name \
                if enq.assigned_to and enq.assigned_to.branch else ""
            )

            ws.cell(row=row, column=7).value = (
                enq.enquiry_source.name
                if enq.enquiry_source else ""
            )

            ws.cell(row=row, column=8).value = (
                enq.status.name
                if enq.status else ""
            )

            ws.cell(row=row, column=9).value = (
                enq.assigned_to.name
                if enq.assigned_to else ""
            )

            ws.cell(row=row, column=10).value = ", ".join(
                c.course_name
                for c in enq.courses.all()
            )

            row += 1

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = (
            'attachment; filename="Enquiry_Report.xlsx"'
        )

        wb.save(response)

        return response

class ExportEnquiryReportPDFAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_enquiry_report_queryset(request)

        response = HttpResponse(
            content_type="application/pdf"
        )

        response["Content-Disposition"] = (
            'attachment; filename="Enquiry_Report.pdf"'
        )

        doc = SimpleDocTemplate(response)

        styles = getSampleStyleSheet()

        elements = []

        elements.append(
            Paragraph(
                "<b>Enquiry Report</b>",
                styles["Title"]
            )
        )

        table_data = [[

            "Date",

            "Enquiry",

            "Candidate",

            "Organization",

            "Branch",

            "Status"

        ]]

        for enq in queryset:

            table_data.append([

                enq.enquiry_date.strftime("%d-%m-%Y"),

                enq.enquiry_code,

                enq.candidate_name,

                enq.assigned_to.organization.name \
                if enq.assigned_to and enq.assigned_to.organization else "",

                enq.assigned_to.branch.name \
                if enq.assigned_to and enq.assigned_to.branch else "",

                enq.status.name
                if enq.status else ""

            ])

        table = Table(table_data)

        table.setStyle(TableStyle([

            ("GRID", (0, 0), (-1, -1), 1, colors.black),

            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),

            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),

            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),

        ]))

        elements.append(table)

        doc.build(elements)

        return response



def validate_followup_report_filters(data):

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")

    if organization_id:

        if not Organization.objects.filter(
            id=organization_id,
            is_active=True
        ).exists():

            raise ValidationError({
                "organizationId":
                    "Selected organization does not exist."
            })

    if branch_id:

        queryset = Branch.objects.filter(
            id=branch_id,
            is_active=True
        )

        if organization_id:

            queryset = queryset.filter(
                organization_id=organization_id
            )

        if not queryset.exists():

            raise ValidationError({
                "branch":
                    "Selected branch does not belong to the selected organization."
            })

def get_followup_report_queryset(request):

    data = request.data

    from_date = data.get("fromDate")
    to_date = data.get("toDate")

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")

    if not from_date or not to_date:

        raise ValidationError({
            "error": "fromDate and toDate are required."
        })

    validate_followup_report_filters(data)

    queryset = EnquiryFollowUp.objects.filter(
        followup_date__range=[from_date, to_date],
        is_active=True
    ).select_related(
        "enquiry",
        "status",
        "assigned_to",
        "assigned_to__organization",
        "assigned_to__branch",
    )

    if organization_id:
        queryset = queryset.filter(
            assigned_to__organization_id=organization_id
        )

    if branch_id:
        queryset = queryset.filter(
            assigned_to__branch_id=branch_id
        )

    return queryset.order_by("-followup_date")

class FollowUpReportAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        try:

            data = request.data

            from_date = data.get("fromDate")
            to_date = data.get("toDate")

            organization_id = data.get("organizationId")
            branch_id = data.get("branch")

            queryset = get_followup_report_queryset(request)

            results = []

            for follow in queryset:

                #emp = follow.enquiry.assigned_to
                emp = follow.assigned_to

                results.append({

                    "id": follow.id,

                    "followup_date": follow.followup_date.strftime("%Y-%m-%d"),

                    "enquiry_code": follow.enquiry.enquiry_code,

                    "candidate_name": follow.enquiry.candidate_name,

                    "mobile": follow.enquiry.mobile_no,

                    "organization": {

                        "id": (
                            emp.organization.id
                            if emp and emp.organization else None
                        ),

                        "name": (
                            emp.organization.name
                            if emp and emp.organization else None
                        )
                    },

                    "branch": {

                        "id": (
                            emp.branch.id
                            if emp and emp.branch else None
                        ),

                        "name": (
                            emp.branch.name
                            if emp and emp.branch else None
                        )
                    },

                    "remark": follow.remark,

                    "status": (
                        follow.status.name
                        if follow.status else "No Status"
                    )
                })

            return Response({

                "metadata": {

                    "report_name": "Follow-up Report",

                    "from_date": from_date,

                    "to_date": to_date,

                    "organization_id": organization_id,

                    "branch_id": branch_id,

                    "generated_at": datetime.now().isoformat()
                },

                "summary": {

                    "total_followups": queryset.count()
                },

                "results": results

            })

        except Exception as e:

            return Response(
                {"error": str(e)},
                status=500
            )

class ExportFollowUpReportExcelAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_followup_report_queryset(request)

        wb = Workbook()

        ws = wb.active

        ws.title = "Follow-up Report"

        headers = [

            "Follow-up Date",

            "Enquiry Code",

            "Candidate",

            "Mobile",

            "Organization",

            "Branch",

            "Status",

            "Remark"
        ]

        for col, header in enumerate(headers, 1):

            cell = ws.cell(row=1, column=col)

            cell.value = header

            cell.font = Font(bold=True)

        row = 2

        for follow in queryset:

            #emp = follow.enquiry.assigned_to
            emp = follow.assigned_to

            ws.cell(row=row, column=1).value = follow.followup_date.strftime("%d-%m-%Y")

            ws.cell(row=row, column=2).value = follow.enquiry.enquiry_code

            ws.cell(row=row, column=3).value = follow.enquiry.candidate_name

            ws.cell(row=row, column=4).value = follow.enquiry.mobile_no

            ws.cell(row=row, column=5).value = (
                emp.organization.name
                if emp and emp.organization else ""
            )

            ws.cell(row=row, column=6).value = (
                emp.branch.name
                if emp and emp.branch else ""
            )

            ws.cell(row=row, column=7).value = (
                follow.status.name
                if follow.status else ""
            )

            ws.cell(row=row, column=8).value = follow.remark

            row += 1

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = 'attachment; filename="FollowUp_Report.xlsx"'

        wb.save(response)

        return response

class ExportFollowUpReportPDFAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_followup_report_queryset(request)

        response = HttpResponse(content_type="application/pdf")

        response["Content-Disposition"] = 'attachment; filename="FollowUp_Report.pdf"'

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(letter)
        )

        styles = getSampleStyleSheet()

        elements = []

        elements.append(
            Paragraph(
                "<b>Follow-up Report</b>",
                styles["Title"]
            )
        )

        table_data = [[

            "Date",

            "Enquiry",

            "Candidate",

            "Organization",

            "Branch",

            "Status",

            "Remark"

        ]]

        for follow in queryset:

            #emp = follow.enquiry.assigned_to
            emp = follow.assigned_to

            table_data.append([

                follow.followup_date.strftime("%d-%m-%Y"),

                follow.enquiry.enquiry_code,

                follow.enquiry.candidate_name,

                (
                    emp.organization.name
                    if emp and emp.organization else ""
                ),

                (
                    emp.branch.name
                    if emp and emp.branch else ""
                ),

                (
                    follow.status.name
                    if follow.status else ""
                ),

                follow.remark or ""

            ])

        table = Table(table_data)

        table.setStyle(TableStyle([

            ("GRID", (0, 0), (-1, -1), 1, colors.black),

            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),

            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),

            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),

            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),

        ]))

        elements.append(table)

        doc.build(elements)

        return response






def validate_attendance_report_filters(data):

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")
    employee_id = data.get("employeeId")

    if organization_id:

        if not Organization.objects.filter(
            id=organization_id,
            is_active=True
        ).exists():

            raise ValidationError({
                "organizationId": "Selected organization does not exist."
            })

    if branch_id:

        queryset = Branch.objects.filter(
            id=branch_id,
            is_active=True
        )

        if organization_id:

            queryset = queryset.filter(
                organization_id=organization_id
            )

        if not queryset.exists():

            raise ValidationError({
                "branch": "Selected branch does not belong to the selected organization."
            })

    if employee_id:

        if not Employee.objects.filter(
            id=employee_id,
            is_active=True
        ).exists():

            raise ValidationError({
                "employeeId": "Selected employee does not exist."
            })

def get_attendance_report_queryset(request):

    data = request.data

    from_date = data.get("fromDate")
    to_date = data.get("toDate")

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")
    employee_id = data.get("employeeId")

    if not from_date or not to_date:

        raise ValidationError({
            "error": "fromDate and toDate are required."
        })

    validate_attendance_report_filters(data)

    queryset = Attendance.objects.filter(
        is_active=True,
        date__range=[from_date, to_date]
    ).select_related(
        "employee",
        "employee__organization",
        "employee__branch"
    )

    if organization_id:

        queryset = queryset.filter(
            employee__organization_id=organization_id
        )

    if branch_id:

        queryset = queryset.filter(
            employee__branch_id=branch_id
        )

    if employee_id:

        queryset = queryset.filter(
            employee_id=employee_id
        )

    return queryset.order_by("-date")

class AttendanceReportAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_attendance_report_queryset(request)

        results = []

        for attendance in queryset:

            results.append({

                "id": attendance.id,

                "date": attendance.date,

                "employee_code": attendance.employee.employee_code,

                "employee_name": attendance.employee.name,

                "organization": {
                    "id": attendance.employee.organization.id if attendance.employee.organization else None,
                    "name": attendance.employee.organization.name if attendance.employee.organization else None
                },

                "branch": {
                    "id": attendance.employee.branch.id if attendance.employee.branch else None,
                    "name": attendance.employee.branch.name if attendance.employee.branch else None
                },

                "status": attendance.status,

                "time_in": attendance.time_in,

                "time_out": attendance.time_out,

                "total_hours": attendance.total_hours,

                "remark": attendance.remark
            })

        return Response({

            "report_metadata": {

                "report_name": "Attendance Report",

                "from_date": request.data.get("fromDate"),

                "to_date": request.data.get("toDate"),

                "organization_id": request.data.get("organizationId"),

                "branch_id": request.data.get("branch"),

                "generated_at": datetime.now().isoformat()
            },

            "summary": {

                "total_records": queryset.count(),

                "present": queryset.filter(status="present").count(),

                "absent": queryset.filter(status="absent").count(),

                "half_day": queryset.filter(status="half_day").count(),

                "on_leave": queryset.filter(status="on_leave").count()
            },

            "results": results

        })

class ExportAttendanceReportExcelAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_attendance_report_queryset(request)

        wb = Workbook()

        ws = wb.active

        ws.title = "Attendance Report"

        headers = [

            "Date",

            "Employee Code",

            "Employee Name",

            "Organization",

            "Branch",

            "Status",

            "Time In",

            "Time Out",

            "Total Hours",

            "Remark"

        ]

        for col, header in enumerate(headers, 1):

            cell = ws.cell(row=1, column=col)

            cell.value = header

            cell.font = Font(bold=True)

        row = 2

        for attendance in queryset:

            emp = attendance.employee

            ws.cell(row=row, column=1).value = attendance.date.strftime("%d-%m-%Y")

            ws.cell(row=row, column=2).value = emp.employee_code

            ws.cell(row=row, column=3).value = emp.name

            ws.cell(row=row, column=4).value = emp.organization.name if emp.organization else ""

            ws.cell(row=row, column=5).value = emp.branch.name if emp.branch else ""

            ws.cell(row=row, column=6).value = attendance.status

            ws.cell(row=row, column=7).value = str(attendance.time_in or "")

            ws.cell(row=row, column=8).value = str(attendance.time_out or "")

            ws.cell(row=row, column=9).value = float(attendance.total_hours or 0)

            ws.cell(row=row, column=10).value = attendance.remark

            row += 1

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = 'attachment; filename="Attendance_Report.xlsx"'

        wb.save(response)

        return response

class ExportAttendanceReportPDFAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_attendance_report_queryset(request)

        response = HttpResponse(content_type="application/pdf")

        response["Content-Disposition"] = 'attachment; filename="Attendance_Report.pdf"'

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(letter)
        )

        styles = getSampleStyleSheet()

        elements = []

        elements.append(
            Paragraph(
                "<b>Attendance Report</b>",
                styles["Title"]
            )
        )

        table_data = [[

            "Date",

            "Employee",

            "Organization",

            "Branch",

            "Status",

            "Time In",

            "Time Out",

            "Hours"

        ]]

        for attendance in queryset:

            emp = attendance.employee

            table_data.append([

                attendance.date.strftime("%d-%m-%Y"),

                emp.name,

                emp.organization.name if emp.organization else "",

                emp.branch.name if emp.branch else "",

                attendance.status,

                str(attendance.time_in or ""),

                str(attendance.time_out or ""),

                str(attendance.total_hours or "")
            ])

        table = Table(table_data)

        table.setStyle(TableStyle([

            ("GRID", (0, 0), (-1, -1), 1, colors.black),

            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),

            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),

            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),

        ]))

        elements.append(table)

        doc.build(elements)

        return response






# class AttendanceReportView(generics.ListAPIView):
#     """
#     Generate JSON report for Attendance
#     """
#     permission_classes = [IsAuthenticated]
#     serializer_class = AttendanceReportSerializer

#     def get_queryset(self):
#         # Start with active records and optimize DB hits with select_related
#         queryset = Attendance.objects.filter(is_active=True).select_related('employee')
        
#         # Grab date strings from the URL query params
#         from_date = self.request.query_params.get('from_date')
#         to_date = self.request.query_params.get('to_date')
#         employee_id = self.request.query_params.get('employee')

#         # Filter by Date Range (Optional)
#         if from_date:
#             queryset = queryset.filter(date__gte=from_date)
#         if to_date:
#             queryset = queryset.filter(date__lte=to_date)
            
#         # Filter by Employee (Optional)
#         if employee_id:
#             queryset = queryset.filter(employee_id=employee_id)

#         return queryset.order_by('-date')



def validate_course_tracker_report_filters(data):

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")
    trainer_id = data.get("trainer")
    status_value = data.get("status")

    if organization_id:

        if not Organization.objects.filter(
            id=organization_id,
            is_active=True
        ).exists():

            raise ValidationError({
                "organizationId":
                    "Selected organization does not exist."
            })

    if branch_id:

        queryset = Branch.objects.filter(
            id=branch_id,
            is_active=True
        )

        if organization_id:

            queryset = queryset.filter(
                organization_id=organization_id
            )

        if not queryset.exists():

            raise ValidationError({
                "branch":
                    "Selected branch does not belong to the selected organization."
            })

    if trainer_id:

        if not Employee.objects.filter(
            id=trainer_id,
            is_active=True
        ).exists():

            raise ValidationError({
                "trainer":
                    "Selected trainer does not exist."
            })

    if status_value:

        valid_status = ["Pending", "Completed", "Cancelled"]

        if status_value not in valid_status:

            raise ValidationError({
                "status":
                    f"Status must be one of {valid_status}"
            })

def get_course_tracker_queryset(request):

    data = request.data

    from_date = data.get("fromDate")
    to_date = data.get("toDate")

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")
    trainer_id = data.get("trainer")
    status_value = data.get("status")

    validate_course_tracker_report_filters(data)

    queryset = CourseTracker.objects.filter(
        is_active=True
    ).select_related(
        "organization",
        "branch",
        "trainer",
        "batch",
        "course"
    )

    if from_date:

        queryset = queryset.filter(
            date__gte=from_date
        )

    if to_date:

        queryset = queryset.filter(
            date__lte=to_date
        )

    if organization_id:

        queryset = queryset.filter(
            organization_id=organization_id
        )

    if branch_id:

        queryset = queryset.filter(
            branch_id=branch_id
        )

    if trainer_id:

        queryset = queryset.filter(
            trainer_id=trainer_id
        )

    if status_value:

        queryset = queryset.filter(
            status__iexact=status_value
        )

    return queryset.order_by("-date")

class CourseTrackerReportAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        try:

            queryset = get_course_tracker_queryset(request)

            serializer = CourseTrackerReportSerializer(
                queryset,
                many=True
            )

            return Response({

                "metadata": {

                    "report_name": "Course Tracker Report",

                    "generated_at": datetime.now().isoformat(),

                    "filters": request.data
                },

                "summary": {

                    "total_records": queryset.count()
                },

                "results": serializer.data

            })

        except Exception as e:

            return Response(
                {
                    "error": str(e)
                },
                status=400
            )

class ExportCourseTrackerExcelAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_course_tracker_queryset(request)

        wb = Workbook()

        ws = wb.active

        ws.title = "Course Tracker Report"

        headers = [

            "Date",
            "Organization",
            "Branch",
            "Trainer",
            "Course",
            "Batch",
            "Status",
            "Remark"

        ]

        for col, header in enumerate(headers, 1):

            cell = ws.cell(row=1, column=col)

            cell.value = header

            cell.font = Font(bold=True)

        row = 2

        for tracker in queryset:

            ws.cell(row=row, column=1).value = tracker.date.strftime("%d-%m-%Y")

            ws.cell(row=row, column=2).value = (
                tracker.organization.name
                if tracker.organization else ""
            )

            ws.cell(row=row, column=3).value = (
                tracker.branch.name
                if tracker.branch else ""
            )

            ws.cell(row=row, column=4).value = (
                tracker.trainer.name
                if tracker.trainer else ""
            )

            ws.cell(row=row, column=5).value = (
                tracker.course.course_name
                if tracker.course else ""
            )

            ws.cell(row=row, column=6).value = (
                tracker.batch.batch_name
                if tracker.batch else ""
            )

            ws.cell(row=row, column=7).value = tracker.status

            ws.cell(row=row, column=8).value = tracker.remark

            row += 1

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = (
            'attachment; filename="Course_Tracker_Report.xlsx"'
        )

        wb.save(response)

        return response

class ExportCourseTrackerPDFAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        queryset = get_course_tracker_queryset(request)

        response = HttpResponse(content_type="application/pdf")

        response["Content-Disposition"] = (
            'attachment; filename="Course_Tracker_Report.pdf"'
        )

        doc = SimpleDocTemplate(response)

        styles = getSampleStyleSheet()

        elements = []

        elements.append(
            Paragraph(
                "<b>Course Tracker Report</b>",
                styles["Title"]
            )
        )

        table_data = [[

            "Date",

            "Organization",

            "Branch",

            "Trainer",

            "Course",

            "Batch",

            "Status"

        ]]

        for tracker in queryset:

            table_data.append([

                tracker.date.strftime("%d-%m-%Y"),

                tracker.organization.name
                if tracker.organization else "",

                tracker.branch.name
                if tracker.branch else "",

                tracker.trainer.name
                if tracker.trainer else "",

                tracker.course.course_name
                if tracker.course else "",

                tracker.batch.batch_name
                if tracker.batch else "",

                tracker.status

            ])

        table = Table(table_data)

        table.setStyle(TableStyle([

            ("GRID", (0, 0), (-1, -1), 1, colors.black),

            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),

            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),

            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),

        ]))

        elements.append(table)

        doc.build(elements)

        return response






def validate_student_attendance_filters(data):

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")

    if organization_id:

        if not Organization.objects.filter(
            id=organization_id,
            is_active=True
        ).exists():

            raise ValidationError({
                "organizationId":
                    "Selected organization does not exist."
            })

    if branch_id:

        queryset = Branch.objects.filter(
            id=branch_id,
            is_active=True
        )

        if organization_id:

            queryset = queryset.filter(
                organization_id=organization_id
            )

        if not queryset.exists():

            raise ValidationError({
                "branch":
                    "Selected branch does not belong to the selected organization."
            })

def get_student_attendance_report_queryset(request):

    data = request.data

    from_date = data.get("fromDate")
    to_date = data.get("toDate")

    organization_id = data.get("organizationId")
    branch_id = data.get("branch")

    if not from_date or not to_date:

        raise ValidationError({
            "error":
                "fromDate and toDate are required."
        })

    validate_student_attendance_filters(data)

    queryset = Admission.objects.filter(
        is_active=True
    ).select_related(
        "organization",
        "branch"
    ).prefetch_related(
        "attendance_records"
    )

    if organization_id:

        queryset = queryset.filter(
            organization_id=organization_id
        )

    if branch_id:

        queryset = queryset.filter(
            branch_id=branch_id
        )

    return (
        queryset,
        from_date,
        to_date,
        organization_id,
        branch_id
    )

class StudentAttendanceReportAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        (
            queryset,
            from_date,
            to_date,
            organization_id,
            branch_id
        ) = get_student_attendance_report_queryset(request)

        results = []

        total_students = 0
        total_present = 0
        total_absent = 0
        total_half_day = 0
        total_leave = 0

        for student in queryset:

            attendance = student.attendance_records.filter(
                is_active=True,
                date__range=[from_date, to_date]
            )

            working_days = attendance.count()

            present = attendance.filter(
                status="present"
            ).count()

            absent = attendance.filter(
                status="absent"
            ).count()

            half_day = attendance.filter(
                status="half_day"
            ).count()

            leave = attendance.filter(
                status="on_leave"
            ).count()

            percentage = 0

            if working_days:

                percentage = round(
                    (
                        present +
                        (half_day * 0.5)
                    ) / working_days * 100,
                    2
                )

            results.append({

                "student_id": student.id,

                "admission_code": student.admission_code,

                "student_name": student.candidate_name,

                "organization": {

                    "id": student.organization.id if student.organization else None,

                    "name": student.organization.name if student.organization else None

                },

                "branch": {

                    "id": student.branch.id if student.branch else None,

                    "name": student.branch.name if student.branch else None

                },

                "working_days": working_days,

                "present_days": present,

                "absent_days": absent,

                "half_days": half_day,

                "leave_days": leave,

                "attendance_percentage": percentage

            })

            total_students += 1
            total_present += present
            total_absent += absent
            total_half_day += half_day
            total_leave += leave

        return Response({

            "metadata": {

                "report_name": "Student Attendance Report",

                "from_date": from_date,

                "to_date": to_date,

                "organization_id": organization_id,

                "branch_id": branch_id,

                "generated_at": datetime.now().isoformat()

            },

            "summary": {

                "total_students": total_students,

                "total_present": total_present,

                "total_absent": total_absent,

                "total_half_days": total_half_day,

                "total_leave_days": total_leave

            },

            "results": results

        })

class ExportStudentAttendanceReportExcelAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        (
            queryset,
            from_date,
            to_date,
            organization_id,
            branch_id
        ) = get_student_attendance_report_queryset(request)

        wb = Workbook()

        ws = wb.active

        ws.title = "Student Attendance Report"

        headers = [

            "Admission No",

            "Student",

            "Organization",

            "Branch",

            "Working Days",

            "Present",

            "Absent",

            "Half Day",

            "Leave",

            "Attendance %"

        ]

        for col, header in enumerate(headers, 1):

            cell = ws.cell(row=1, column=col)

            cell.value = header

            cell.font = Font(bold=True)

        row = 2

        for student in queryset:

            attendance = student.attendance_records.filter(
                is_active=True,
                date__range=[from_date, to_date]
            )

            working = attendance.count()

            present = attendance.filter(status="present").count()

            absent = attendance.filter(status="absent").count()

            half = attendance.filter(status="half_day").count()

            leave = attendance.filter(status="on_leave").count()

            percentage = 0

            if working:

                percentage = round(
                    (
                        present +
                        (half * 0.5)
                    ) / working * 100,
                    2
                )

            ws.cell(row=row, column=1).value = student.admission_code
            ws.cell(row=row, column=2).value = student.candidate_name
            ws.cell(row=row, column=3).value = student.organization.name if student.organization else ""
            ws.cell(row=row, column=4).value = student.branch.name if student.branch else ""
            ws.cell(row=row, column=5).value = working
            ws.cell(row=row, column=6).value = present
            ws.cell(row=row, column=7).value = absent
            ws.cell(row=row, column=8).value = half
            ws.cell(row=row, column=9).value = leave
            ws.cell(row=row, column=10).value = percentage

            row += 1

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = 'attachment; filename="Student_Attendance_Report.xlsx"'

        wb.save(response)

        return response

class ExportStudentAttendanceReportPDFAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        (
            queryset,
            from_date,
            to_date,
            organization_id,
            branch_id
        ) = get_student_attendance_report_queryset(request)

        response = HttpResponse(content_type="application/pdf")

        response["Content-Disposition"] = 'attachment; filename="Student_Attendance_Report.pdf"'

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(letter)
        )

        styles = getSampleStyleSheet()

        elements = []

        elements.append(
            Paragraph(
                "<b>Student Attendance Report</b>",
                styles["Title"]
            )
        )

        table_data = [[

            "Admission",

            "Student",

            "Organization",

            "Branch",

            "Working",

            "Present",

            "Absent",

            "Half",

            "Leave",

            "%"

        ]]

        for student in queryset:

            attendance = student.attendance_records.filter(
                is_active=True,
                date__range=[from_date, to_date]
            )

            working = attendance.count()

            present = attendance.filter(status="present").count()

            absent = attendance.filter(status="absent").count()

            half = attendance.filter(status="half_day").count()

            leave = attendance.filter(status="on_leave").count()

            percentage = 0

            if working:

                percentage = round(
                    (
                        present +
                        (half * 0.5)
                    ) / working * 100,
                    2
                )

            table_data.append([

                student.admission_code,

                student.candidate_name,

                student.organization.name if student.organization else "",

                student.branch.name if student.branch else "",

                working,

                present,

                absent,

                half,

                leave,

                f"{percentage}%"

            ])

        table = Table(table_data)

        table.setStyle(TableStyle([

            ("GRID", (0, 0), (-1, -1), 1, colors.black),

            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),

            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),

            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("BACKGROUND", (0, 1), (-1, -1), colors.beige)

        ]))

        elements.append(table)

        doc.build(elements)

        return response




        
